from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .config import EXECUTION_SHEET, PROFILE_GLOBAL_SHEET, PROFILE_ORIGIN_SHEET
from .models import ProfileExecution

_HEADER_FILL = PatternFill("solid", fgColor="163A5F")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _safe_excel_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        try:
            return value.item()
        except (ValueError, TypeError):
            pass
    return value


def _unique_table_name(workbook, preferred: str) -> str:
    existing = {
        table.name
        for worksheet in workbook.worksheets
        for table in worksheet.tables.values()
    }
    if preferred not in existing:
        return preferred
    counter = 2
    while f"{preferred}{counter}" in existing:
        counter += 1
    return f"{preferred}{counter}"


def _write_dataframe_sheet(workbook, sheet_name: str, dataframe: pd.DataFrame) -> None:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False

    for col_index, column_name in enumerate(dataframe.columns, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=str(column_name))
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row in enumerate(dataframe.itertuples(index=False, name=None), start=2):
        for col_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=col_index, value=_safe_excel_value(value))

    percentage_columns = {
        "% Valores únicos",
        "%Nulos",
        "%Espacios",
    }
    for col_index, column_name in enumerate(dataframe.columns, start=1):
        if column_name in percentage_columns:
            for row_index in range(2, len(dataframe) + 2):
                worksheet.cell(row=row_index, column=col_index).number_format = '0.00"%"'

    if not dataframe.empty:
        end_cell = f"{get_column_letter(len(dataframe.columns))}{len(dataframe) + 1}"
        table = Table(
            displayName=_unique_table_name(workbook, f"tbl_{sheet_name}"),
            ref=f"A1:{end_cell}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    for col_index, column_name in enumerate(dataframe.columns, start=1):
        sample_values = [str(column_name)] + [
            "" if value is None else str(value)
            for value in dataframe.iloc[:200, col_index - 1].tolist()
        ]
        width = min(max(max(map(len, sample_values)) + 2, 12), 34)
        worksheet.column_dimensions[get_column_letter(col_index)].width = width

    worksheet.row_dimensions[1].height = 32


def _write_execution_sheet(workbook, execution: ProfileExecution) -> None:
    if EXECUTION_SHEET in workbook.sheetnames:
        del workbook[EXECUTION_SHEET]
    worksheet = workbook.create_sheet(EXECUTION_SHEET)
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    rows = [
        ("Indicador", "Valor"),
        ("Archivo fuente", execution.source_file),
        ("Hoja", execution.sheet_name),
        ("Tabla o rango", execution.table_name),
        ("Referencia", execution.table_reference),
        ("Columna de segmentación", execution.segment_column),
        ("Cantidad de registros", execution.row_count),
        ("Cantidad de campos", execution.column_count),
        ("Sistemas de origen detectados", execution.origin_count),
        ("Fecha y hora de ejecución", execution.executed_at),
        ("Tiempo de ejecución (segundos)", execution.elapsed_seconds),
        ("Definición % Valores únicos", "Distintos no nulos / registros no nulos"),
        ("Definición %Nulos", "Nulos físicos y marcadores configurados / registros"),
        ("Definición %Espacios", "Registros con espacios iniciales, finales o repetidos / registros"),
        ("Definición Llave", "Candidata técnica: sin nulos y 100% única dentro del alcance"),
    ]

    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=col_index, value=value)

    for cell in worksheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    worksheet.column_dimensions["A"].width = 38
    worksheet.column_dimensions["B"].width = 80
    for row in worksheet.iter_rows(min_row=2, max_col=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def build_profiled_workbook(
    source_file: Path,
    global_profile: pd.DataFrame,
    segmented_profile: pd.DataFrame,
    execution: ProfileExecution,
) -> bytes:
    """Genera una copia del libro original con tres hojas nuevas de resultados."""
    keep_vba = source_file.suffix.lower() == ".xlsm"
    workbook = load_workbook(source_file, keep_vba=keep_vba, data_only=False)
    try:
        _write_dataframe_sheet(workbook, PROFILE_GLOBAL_SHEET, global_profile)
        _write_dataframe_sheet(workbook, PROFILE_ORIGIN_SHEET, segmented_profile)
        _write_execution_sheet(workbook, execution)
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
    finally:
        workbook.close()
