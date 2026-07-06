from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from .config import DATA_DIRECTORY, SUPPORTED_EXCEL_EXTENSIONS
from .models import DatasetSelection


@dataclass(frozen=True, slots=True)
class ExcelTableInfo:
    """Metadatos de una tabla o rango dentro de una hoja."""

    name: str
    reference: str
    is_excel_table: bool

    @property
    def label(self) -> str:
        source_type = "Tabla de Excel" if self.is_excel_table else "Rango detectado"
        return f"{self.name} · {self.reference} · {source_type}"


def list_excel_files(directory: Path = DATA_DIRECTORY) -> list[Path]:
    """Lista archivos Excel válidos"""
    directory.mkdir(parents=True, exist_ok=True)
    return sorted(
        (
            path
            for path in directory.iterdir()
            if path.is_file()
            and path.suffix.lower() in SUPPORTED_EXCEL_EXTENSIONS
            and not path.name.startswith("~$")
        ),
        key=lambda path: path.name.casefold(),
    )


def list_sheet_names(file_path: Path) -> list[str]:
    """Devuelve los nombres de las hojas sin cargar las fórmulas para cálculo."""
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _used_value_range(worksheet) -> str | None:
    """Calcula el rango real con valores y evita depender solo del formato aplicado."""
    min_row = min_col = None
    max_row = max_col = None

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            min_row = cell.row if min_row is None else min(min_row, cell.row)
            max_row = cell.row if max_row is None else max(max_row, cell.row)
            min_col = cell.column if min_col is None else min(min_col, cell.column)
            max_col = cell.column if max_col is None else max(max_col, cell.column)

    if None in (min_row, min_col, max_row, max_col):
        return None

    from openpyxl.utils import get_column_letter

    return (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )


def list_tables(file_path: Path, sheet_name: str) -> list[ExcelTableInfo]:
    """Detecta tablas, si no existen detecta rangos con datos."""
    workbook = load_workbook(file_path, read_only=False, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        tables = [
            ExcelTableInfo(name=table.name, reference=table.ref, is_excel_table=True)
            for table in worksheet.tables.values()
        ]
        if tables:
            return sorted(tables, key=lambda item: item.name.casefold())

        used_range = _used_value_range(worksheet)
        if used_range is None:
            return []

        return [
            ExcelTableInfo(
                name=f"RANGO_{sheet_name}",
                reference=used_range,
                is_excel_table=False,
            )
        ]
    finally:
        workbook.close()


def _deduplicate_headers(values: Iterable[object]) -> list[str]:
    """Normaliza encabezados vacíos o repetidos para construir un DataFrame válido."""
    headers: list[str] = []
    counts: dict[str, int] = {}

    for index, value in enumerate(values, start=1):
        base = str(value).strip() if value is not None else ""
        base = base or f"Columna_{index}"
        counts[base] = counts.get(base, 0) + 1
        suffix = counts[base]
        headers.append(base if suffix == 1 else f"{base}_{suffix}")

    return headers


def read_selection(selection: DatasetSelection) -> pd.DataFrame:
    """Lee únicamente el rango seleccionado y devuelve sus datos como DataFrame."""
    workbook = load_workbook(
        selection.file_path,
        read_only=False,
        data_only=True,
        keep_vba=selection.file_path.suffix.lower() == ".xlsm",
    )
    try:
        worksheet = workbook[selection.sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(selection.table_reference)
        rows = list(
            worksheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
        )
    finally:
        workbook.close()

    if not rows:
        raise ValueError("El rango seleccionado no contiene datos.")

    headers = _deduplicate_headers(rows[0])
    dataframe = pd.DataFrame(rows[1:], columns=headers)
    dataframe = dataframe.dropna(axis=0, how="all")

    if dataframe.empty:
        raise ValueError("La tabla solo contiene encabezados; no hay registros para perfilar.")

    return dataframe.reset_index(drop=True)
